# -*- coding: utf-8 -*-

################### どこぞの高専の席替え用プログラム2022  ##################
# 使いかた
# 1. python環境を入れる
# 2. ターミナルで  pip install openpyxl pandas する
# 3. ターミナルで個のファイルを実行する
#     例 python seating.py      -> シャッフルされた座席ファイルが出来る
#     例 python seating.py aaa  -> 出席番号順の席ファイルが出来る
#########################################################################


import openpyxl          # Excelの操作に使う
import pandas as pd      # 名標からのデータ読み込みに使う
import datetime          # 日付に使う
from openpyxl.styles import PatternFill  # 性別によって座席表を色付けするのに使う
import sys    # コマンドライン引数の処理に使う


# 読み込むファイル. 名標と座席のフォーマットあり.
EXCEL_FILE = "2022座席.xlsx"
df = pd.read_excel(EXCEL_FILE, sheet_name="account")

# ４月の初め用
if len(sys.argv) == 2:  # 引数ありならそのまま
    df_s = df
else:
    df_s = df.sample(frac=1).reset_index(drop=True)  # 引数なしなら読み込んだ名標データをシャッフル

# 操作用のシート
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.get_sheet_by_name(name="master")

redFill = PatternFill(start_color='FFFF0080',  # 女子の場合の色の設定
                      end_color='FFFF0080',
                      fill_type='solid'
                      )


idx = 0  # カウンタを設定して
for j in range(2, 24, 3):  # ループに入れ
    for i in reversed(range(5, 21, 3)):  # Excelのセルのインデックスで回していく
        if (j < 19) and i == 5:
            continue
        account = df_s.iloc[idx]  # 現在処理中のデータ
        ws.cell(i, j).value = str(account['番号'])          # 出席番号の設定
        ws.cell(i, j + 1).value = str(account['フリガナ'])  # ふりがなの設定
        ws.cell(i + 1, j + 1).value = str(account['氏名'])  # 名前の設定
        if account['性'] == 1:  # 女子なら
            ws.cell(i + 1, j).fill = redFill  # 色付け

        idx = idx + 1  # カウンタを更新して
        # ループにもどれ

today = datetime.date.today()  # TODAYに日付をセット
ws["W25"].value = str(today)   # 日付の書き込み


wb.save(f"座席_{str(today)}.xlsx")  # 保存
